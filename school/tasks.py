import json
import os
import traceback
from itertools import chain
from os import path
import requests
import openpyxl
from background_task import background
from django.db.models import F
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

from mylib.queryset2excel import importExcelCsv, serializerToRow, get_header_column_name, serializerErrorToRow
from region.models import SubCounty, County
from school.models import SchoolsStudentsImport, School, Stream, Student

from school.schools_students_imports.serializers import StudentSchoolSerializer, SchoolImportSerializer
from school.student.serializers import StudentSerializer

from django.conf import settings
import tempfile
from django.core.files.storage import default_storage

# from iterable_orm import QuerySet


def get_county_id(valid_data, entity=""):
    county = valid_data.get("{}county".format(entity), None)
    if county == None:
        return None

    queryset = County.objects.filter(name__iexact=county)
    if queryset.exists():
        return queryset[0].id
    c = County.objects.create(name=county)
    return c.id


def is_value_empty(value):
    if value == None:
        return True
    if f"{value}".strip() == "":
        return True
    return False


def get_subcounty_id(valid_data, entity=""):
    subcounty = valid_data.get("{}subcounty".format(entity), None)
    if subcounty == None:
        return None
    queryset = SubCounty.objects.filter(name__iexact=subcounty)
    if queryset.exists():
        return queryset[0].id
    county_id = get_county_id(valid_data, entity)

    if county_id == None:
        return None

    sc = SubCounty.objects.create(county_id=county_id, name=subcounty)
    return sc.id


def get_school_from_nemis(valid_data):
    nemis_code = valid_data.get("school_nemis_code").upper()
    school = valid_data.get("school")
    school_ownership = valid_data.get("school_ownership")
    # SchoolImportSerializer

    if School.objects.filter(emis_code__iexact=nemis_code).exists():
        # print("Got school nemis")
        return School.objects.filter(emis_code__iexact=nemis_code).first().id

    # if School.objects.filter(name__iexact=school).exists():
    #     print("Got school name")

    #     return School.objects.filter(name__iexact=school).first().id

    ## SubCounty
    subcounty = get_subcounty_id(valid_data)

    sch = School.objects.create(sub_county_id=subcounty, name=school, ownership=school_ownership, emis_code=nemis_code)
    return sch.id


def get_stream_name(valid_data):
    stream = valid_data.get("stream")
    extractedints = list(filter(str.isdigit, stream))
    if len(extractedints) < 1:
        return None
    stream = stream[stream.index(extractedints[0]) + 1 :]
    base_class = extractedints[0]
    return (base_class, stream)


def get_stream_id(valid_data):
    stream = valid_data.get("stream")
    school = get_school_from_nemis(valid_data)

    extractedints = list(filter(str.isdigit, stream))
    if len(extractedints) < 1:
        return None
    st_name = get_stream_name(valid_data)

    stream = st_name[1]
    base_class = st_name[0]

    # print(f"Stream {stream}  Base class {base_class} ")

    streams = list(Stream.objects.filter(school_id=school, name=stream, base_class=base_class))
    try:
        if len(streams) < 1:
            sts = Stream.objects.create(name=stream, school_id=school, base_class=base_class)
            return sts.id
        else:
            return streams[0].id
    except Exception as e:
        print(e)
        return None


def get_stud_filters(first_name=None, middle_name=None, last_name=None, upi=None, school_nemis_code=None, stream_name=None, base_class=None):
    queryFilters = {}
    if first_name:
        queryFilters["first_name__icontains"] = first_name

    if middle_name:
        queryFilters["middle_name__icontains"] = middle_name

    if last_name:
        queryFilters["last_name__icontains"] = last_name

    if upi:
        queryFilters["upi"] = upi

    if school_nemis_code != None:
        queryFilters["stream__school__emis_code__icontains"] = school_nemis_code

    if stream_name != None and stream_name != "":
        queryFilters["stream__name__icontains"] = stream_name

    if base_class != None:
        queryFilters["stream__base_class"] = base_class
    return queryFilters


def get_guardian_name(valid_data):
    father_name = valid_data.get("father_name")
    mother_name = valid_data.get("father_name")
    guardian_name = valid_data.get("father_name")
    if father_name != None and "N/A" not in father_name.upper():
        return father_name
    if mother_name != None and "N/A" not in mother_name.upper():
        return mother_name
    return guardian_name


def get_student_id(valid_data, import_instance=None):
    # pass
    ## Ccounty and sub_Cunty
    sub_county_id = get_subcounty_id(valid_data)

    learner_sub_county_id = get_subcounty_id(valid_data, "learner_")

    # print(learner_sub_county_id)
    # print(valid_data)
    # print(valid_data.get("{}subcounty".format("learner_"),None))

    ## Confirm School Existst
    school_id = get_school_from_nemis(valid_data)
    school = School.objects.get(id=school_id)
    print("SCHOOL")
    print(valid_data.get("school_nemis_code"))
    print(school.name)
    print(school.emis_code)

    ## Confirm Stream Exists
    stream_id = get_stream_id(valid_data)
    # print(f"Got the strema st{stream_id}")

    if stream_id == None:
        return {"valid": False, "data": None, "error": "No Valid class provided.(1-8)"}

    ## Create Student
    stud_data = {
        "stream": stream_id,
        "gender": valid_data.get("gender"),
        "sub_county": learner_sub_county_id,
        "guardian_sub_county": learner_sub_county_id,
        "status": valid_data.get("status"),
        "first_name": valid_data.get("first_name"),
        "middle_name": valid_data.get("middle_name"),
        "last_name": valid_data.get("last_name"),
        "guardian_name": get_guardian_name(valid_data),
        "guardian_phone": valid_data.get("guardian_phone"),
        "guardian_email": valid_data.get("guardian_email"),
        "distance_from_school": valid_data.get("distance_from_school"),
        "upi": valid_data.get("upi"),
        "village": valid_data.get("village"),
        "admission_no": valid_data.get("admission_number"),
    }
    if valid_data.get("date_enrolled"):
        stud_data["date_enrolled"] = valid_data.get("date_enrolled").date()

    if valid_data.get("date_of_birth"):
        stud_data["date_of_birth"] = valid_data.get("date_of_birth").date()

    st = StudentSerializer(data=stud_data)
    if not st.is_valid():
        # print(st.errors)
        return {"valid": False, "data": None, "error": str(st.errors)}

    valid_stud = st.validated_data
    school_nemis_code = valid_data.get("school_nemis_code").upper()
    # print(school_nemis_code)
    st_name = get_stream_name(valid_data)
    stream = st_name[1]
    base_class = st_name[0]

    queryFilters = get_stud_filters(
        first_name=valid_stud.get("first_name"),
        last_name=valid_stud.get("last_name"),
        middle_name=valid_stud.get("middle_name"),
        upi=valid_stud.get("upi"),
        school_nemis_code=school_nemis_code,
        stream_name=stream,
        base_class=base_class,
    )

    # print(queryFilters)

    ## Check upi first
    if "upi" in queryFilters:
        queryset = Student.objects.filter(upi__iexact=queryFilters.get("upi"))
        if queryset.exists():
            update_learner(valid_data, queryset.first(), import_instance)
            return {"valid": True, "data": None}

    # print()
    queryset = Student.objects.filter(**queryFilters)

    if queryset.exists():
        update_learner(valid_data, queryset.first(), import_instance)
        return {"valid": True, "data": None}

    print(queryFilters)
    return {"valid": True, "data": valid_stud}


def update_learner(valid_data, student, import_instance=None):
    # print("SHould update...")
    if import_instance == None:
        return

    # print("Import intsance found...")

    status = valid_data.get("status")
    try:
        if import_instance.update_learner:
            # print("Updating learner status...")

            student.status = status
            student.save()

    except Exception as e:
        print(e)
        pass


def append_import_count(import_id, rows_count):
    import_task = SchoolsStudentsImport.objects.get(id=import_id)
    import_task.append_count(rows_count)


def set_duplicates_count(import_id, count, total_count):
    import_task = SchoolsStudentsImport.objects.get(id=import_id)
    import_task.duplicates_count = count
    import_task.save()


def nemis_to_daa(nemis):
    data = {
        "school": nemis["Institution_Name"],
        "school_nemis_code": nemis["Institution_Code"],
        "county": nemis["County_Name"],
        "county_nemis_code": nemis["County_Code"],
        "subcounty": nemis["Sub_County_Name"],
        "first_name": nemis["FirstName"],
        "middle_name": nemis["OtherNames"],
        "last_name": nemis["Surname"],
        "gender": nemis["Gender"],
        "stream": nemis["Class_Name"],
        "father_name": nemis["Father_Name"],
        "mother_name": nemis["Mother_Name"],
        "guardian_name": nemis["Guardian_Name"],
        "learner_county": nemis["County_Name"],
        "learner_subcounty": nemis["Sub_County_Name"],
        "date_enrolled": nemis["Date_Captured"],
        "guardian_phone": nemis["Father_Contacts"],
        "upi": nemis["UPI"],
        "date_of_birth": nemis["DOB"],
    }
    if data["guardian_name"] == "" and data["father_name"] != "":
        data["guardian_name"] = data["father_name"]

    if data["guardian_name"] == "" and data["mother_name"] != "":
        data["guardian_name"] = data["mother_name"]
    return data


def get_nemis_data(url):
    r = requests.get(url)
    # print(r.text)
    if r.status_code == 200:
        return json.loads(r.text)
    return []
    # count = 0
    # for raw_rsvp in r.iter_lines():
    #     count += 1
    #     print(count)
    #     yield json.loads(raw_rsvp)


NEMIS_DATA_URL = ""


NEMIS_GROUPING_URL = {
}


def ensure_dir_or_create(dir):
    if not os.path.exists(dir):
        os.makedirs(dir)


@background(schedule=1,queue='nemis-import')
def import_nemis_data(import_id):
    import_task = SchoolsStudentsImport.objects.get(id=import_id)
    import_task.prepare_import()

    dulicates_count = 0
    errors_count = 0
    processed_row_count = 0
    row_info = {}
    rows_count = 0

    errors_wb = Workbook(write_only=True)
    errors_ws = errors_wb.create_sheet()

    ser = StudentSchoolSerializer()
    fields = ser.get_fields()
    required_fields = [f for f in fields if fields[f].required]
    error_headers = [f for f in fields] + ["error_description"]

    errors_ws.append([get_header_column_name(header) for header in error_headers])

    def set_duplicate_with_count():
        SchoolsStudentsImport.objects.filter(id=import_id).update(duplicates_count=F("duplicates_count") + 1)

    def finish():
        import_task = SchoolsStudentsImport.objects.get(id=import_id)
        import_task.finish()

    def set_row_errors(row, row_number, error_description):
        SchoolsStudentsImport.objects.filter(id=import_id).update(error_rows_count=F("error_rows_count") + 1)
        row["error_description"] = str(error_description)
        try:
            errors_ws.append(serializerErrorToRow(error_headers, row))
        except Exception as e:
            print("Ws Appen", e)

    def create_new_student(new_student):
        resp = Student.objects.create(**new_student)
        # print(resp)
        SchoolsStudentsImport.objects.filter(id=import_id).update(new_students_created=F("new_students_created") + 1)
        return

    def update_import_task_name(row):
        import_task = SchoolsStudentsImport.objects.get(id=import_id)
        value_field = "County_Name" if import_task == "NMC" else "Sub_County_Name"
        import_task.name = f"{import_task.get_import_type_display()} {row[value_field]}"
        print(import_task.name)
        import_task.save()

    def saveErroFile(error_rows):
        print(f"Saving {error_rows} errors")
        if error_rows > 0:
            print("\n\n\n******SAVING ERRORS*****\n\n")

            print("No erros")
            errors_dir = path.join("Imports")

            # ensure_dir_or_create()
            file_path = path.join("imports", "Import-{}-Errors.xlsx".format(import_id))
            file_name = "temporary1.xlsx"
            errors_wb.save(file_name)

            res = default_storage.save(file_path, open(file_name, "rb"))
            print(f"Default storage returned {res}")
            xp = SchoolsStudentsImport.objects.get(id=import_id)

            xp.finish(errors_file_path=res)
            try:
                os.remove(file_name)
            except Exception as e:
                print(e)
        else:
            print("No error to save")

    def process_single_row(row_info):
        global processed_row_count
        # print(dulicates_count)
        global errors_count
        try:
            row = row_info
            ## Apend default status as previously enrolled
            row["status"] = "PE"
            ser = StudentSchoolSerializer(data=row)

            if ser.is_valid():
                valid_data = ser.validated_data
                new_student_info = get_student_id(valid_data=valid_data)
                # print(new_student_info)
                if not new_student_info["valid"]:
                    print("INvalid studend")
                    error = new_student_info.get("error", "Unknow Error")
                    set_row_errors(row, 0, error)
                elif new_student_info["valid"]:
                    # print("Vald..")
                    new_student = new_student_info["data"]
                    if new_student != None:
                        create_new_student(new_student)
                    else:
                        set_duplicate_with_count()
            else:
                serialiserErrors = {}
                for error in ser.errors:
                    serialiserErrors[error] = list(map(lambda erro: str(erro), ser.errors[error]))
                errorsStringArray = list(map(lambda key: "{} - {}".format(get_header_column_name(key), ", ".join(serialiserErrors[key])), serialiserErrors))
                #
                errorsString = " \n".join(errorsStringArray)
                print("Failed to validate row")
                # print(row_info)
                # print(errorsString)
                set_row_errors(row, 0, errorsString)

            append_import_count(import_id, 1)
        except Exception as e:
            print("Failed to do anything...")
            print(e)
            # set_row_errors(row, processed_row_count, e)
            pass

    if import_task.import_type == "JS":
        # Check if it's an array
        data = import_task.raw_data
        import_task.start(len(data))
        for d in data:
            processed_row_count += 1
            process_single_row(nemis_to_daa(d))
    else:
        url = f"{NEMIS_DATA_URL}/{NEMIS_GROUPING_URL[import_task.import_type]}/{import_task.nemis_group_id}/{import_task.nemis_institution_level}"
        print(url)
        try:
            data = get_nemis_data(url)
            # import_task.start(rows_count)
            print("hello")
            print(len(data))
            import_task.start(len(data))
            count = 0
            for row in data:
                if count == 0:
                    update_import_task_name(row)
                count += 1
                process_single_row(nemis_to_daa(row))

        except Exception as e:
            import_task = SchoolsStudentsImport.objects.get(id=import_id)
            import_task.set_errors(str(e))
            pass

    if errors_count > 0:
        saveErroFile(errors_count)
    else:
        finish()

    # import_task.prepare_import()
    # import_task.start(rows_count)
    # import_task.finish(error_rows_count=errors_count)
    # Handle getting of the jsons and converign to the required stuff


# def set_errors()


@background(schedule=1,queue='file-import')
def import_school_students(import_id):
    print("JOB\tImport:#{}".format(import_id))

    if not openpyxl.xml.lxml_available():
        print("No lxml installed")
        SchoolsStudentsImport.objects.filter(id=import_id).update(step="F", errors="lxml not installed.")
        return

    ## Create an errors list Excel
    errors_wb = Workbook(write_only=True)
    errors_ws = errors_wb.create_sheet()
    error_rows = 0
    global students_to_create
    students_to_create = []
    # students_to_create
    try:
        import_task = SchoolsStudentsImport.objects.get(id=import_id)

        print(import_task.import_file.name)
        cloud_file = default_storage.open(import_task.import_file.name)
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.write(cloud_file.read())
        cloud_file.close()
        temp_file.close()

        print(temp_file.name)
        # print(temp_file.file)
        # file_path = path.join(MEDIA_ROOT, import_task.import_file.name)
        # print(file_path)
        import_task.prepare_import()

        file_path = temp_file.name

        sheets_info_gen = importExcelCsv(filename=file_path, headers_only=True, include_rows_count=True)
        sheets = []
        for h in sheets_info_gen:
            sheets = h
        # print(sheets)

        if len(sheets) < 1:
            import_task = SchoolsStudentsImport.objects.get(id=import_id)
            import_task.set_errors("No headers found")
            return

        sheet = sheets[0]
        rows_count = sheet["rows_count"]
        headers = sheet["headers"]

        # print(rows_count)

        import_task = SchoolsStudentsImport.objects.get(id=import_id)
        import_task.start(rows_count)

        ##Chek the first line
        ser = StudentSchoolSerializer()
        fields = ser.get_fields()
        required_fields = [f for f in fields if fields[f].required]

        missing_colums = []
        for required_field in required_fields:
            if required_field not in headers:
                missing_colums.append(required_field)

        if len(missing_colums) > 0:
            import_task = SchoolsStudentsImport.objects.get(id=import_id)
            import_task.set_errors("The following columns are required. {}".format(", ".join([get_header_column_name(header) for header in missing_colums])))
            print(f"Missing columns {missing_colums}")
            return

        def get_row_number_title():
            return "row_number (Import #{})".format(import_id)

        ## Set up the errors file headers
        error_headers = [get_row_number_title()] + [f for f in fields] + ["error_description"]
        errors_ws.append([get_header_column_name(header) for header in error_headers])

        rowcount = 1
        global dulicates_count
        dulicates_count = 0
        global processed_row_count
        processed_row_count = 0
        new_students = 0

        def set_row_errors(row, row_number, error_description):
            row[get_row_number_title()] = row_number + 1
            row["error_description"] = str(error_description)
            try:
                # print(serializerErrorToRow(error_headers, row))
                errors_ws.append(serializerErrorToRow(error_headers, row))
            except Exception as e:
                print("Ws Appen", e)

        def bulk_create_students():
            global students_to_create
            res_count = len(Student.objects.bulk_create(students_to_create))
            print(res_count)
            students_to_create = []

            import_task = SchoolsStudentsImport.objects.get(id=import_id)
            import_task.new_students_created += res_count
            import_task.save()

        def create_new_student(new_student):
            global students_to_create
            global students_to_create_queryset
            stud = Student.objects.create(**new_student)
            print(stud.stream.school.name)
            print(stud.stream.school.emis_code)
            import_task = SchoolsStudentsImport.objects.get(id=import_id)
            import_task.new_students_created += 1
            import_task.save()
            return

        def set_duplicate_with_count():
            global dulicates_count
            global processed_row_count
            dulicates_count += 1
            set_row_errors(row, processed_row_count, "Duplicate")

        # print("Looping dem rows")
        for row_info in importExcelCsv(filename=file_path):
            if processed_row_count >= rows_count:
                continue

            processed_row_count += 1

            try:
                if row_info["header_row"]:
                    processed_row_count -= 1
                    continue
                row = row_info["row"]

                # Ignore empty rows
                keys = list(row.keys())

                empty_keys = [key for key in keys if is_value_empty(row[key])]
                if len(keys) == len(empty_keys):
                    processed_row_count -= 1
                    continue

                # print("Gettin here")
                # error_fields_to_reset_to_null = ["date_of_birth", "date_enrolled"]
                # ser1 = StudentSchoolSerializer(data=row)
                # if not ser1.is_valid():
                #     errors_fields = ser1.errors.keys()
                #     print(errors_fields)

                ser = StudentSchoolSerializer(data=row)
                # print("Checking serializer")
                if ser.is_valid():
                    # print("Serializer Valid !")
                    # print(ser.validated_data)
                    valid_data = ser.validated_data
                    new_student_info = get_student_id(valid_data=valid_data, import_instance=import_task)

                    if not new_student_info["valid"]:
                        error = new_student_info.get("error", "Unknow Error")
                        error_rows += 1
                        set_row_errors(row, processed_row_count, error)

                    elif new_student_info["valid"]:
                        new_student = new_student_info["data"]
                        if new_student != None:
                            print("Creating students..")
                            if import_task.should_import:
                                create_new_student(new_student)
                        else:
                            print("Duplicate..")

                            set_duplicate_with_count()
                else:
                    print("Invalid !")
                    error_rows += 1
                    serialiserErrors = {}

                    for error in ser.errors:
                        serialiserErrors[error] = list(map(lambda erro: str(erro), ser.errors[error]))
                    errorsStringArray = list(map(lambda key: "{} - {}".format(get_header_column_name(key), ", ".join(serialiserErrors[key])), serialiserErrors))
                    #
                    errorsString = " \n".join(errorsStringArray)
                    # print(errorsString)
                    # print(row)
                    set_row_errors(row, processed_row_count, errorsString)

                append_import_count(import_id, 1)
            except Exception as e:
                print("The error is here")
                print(e)
                error_rows += 1
                set_row_errors(row, processed_row_count, e)
                pass

        # Create any remaining students.
        # create_new_student(None)

        ## If
        set_duplicates_count(import_id, dulicates_count, processed_row_count)
        import_task = SchoolsStudentsImport.objects.get(id=import_id)

        if error_rows > 0:
            print("Yeah erros")
            file_path = path.join("imports", "Import-{}-Errors.xlsx".format(import_id))
            file_name = "temporary1.xlsx"
            errors_wb.save(file_name)
            res = default_storage.save(file_path, open(file_name, "rb"))

            xp = SchoolsStudentsImport.objects.get(id=import_id)
            xp.finish(errors_file_path=res, error_rows_count=error_rows)

            try:
                os.remove(file_name)
            except Exception as e:
                print(e)
        else:
            file_path = path.join("imports", "Import-{}-Errors.xlsx".format(import_id))
            file_name = "temporary1.xlsx"
            errors_wb.save(file_name)
            try:
                os.remove(file_name)
            except Exception as e:
                print(e)
            import_task.finish()

        print("Done", rows_count, import_task.imported_rows_count, "<-Add +->", import_task.duplicates_count, import_task.new_students_created, error_rows)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(e)

    except InvalidFileException as ife:
        print("Error IFE")
        import_task = SchoolsStudentsImport.objects.get(id=import_id)
        import_task.set_errors(ife.args[0])
        # print(import_task.errors)
        # print()
        if error_rows > 0:
            file_path = path.join("imports", "Import-{}-Errors.xlsx".format(import_id))
            file_name = "temporary2.xlsx"
            errors_wb.save(file_name)
            res = default_storage.save(file_path, open(file_name, "rb"))
            errors_wb.save(file_path)

            xp = SchoolsStudentsImport.objects.get(id=import_id)
            xp.finish(errors_file_path=res, error_rows_count=error_rows)

            try:
                os.remove(file_name)
            except Exception as e:
                print(e)
        else:
            import_task.finish()

    except Exception as e:
        import_task = SchoolsStudentsImport.objects.get(id=import_id)
        import_task.set_errors("Failed")
        print(traceback.format_exc())

        if error_rows > 0:
            file_path = path.join("imports", "Import-{}-Errors.xlsx".format(import_id))
            file_name = "temporary3.xlsx"
            errors_wb.save(file_name)
            res = default_storage.save(file_path, open(file_name, "rb"))
            errors_wb.save(file_path)

            xp = SchoolsStudentsImport.objects.get(id=import_id)
            xp.finish(errors_file_path=res, error_rows_count=error_rows)

            try:
                os.remove(file_name)
            except Exception as e:
                print(e)
        else:
            import_task.finish()

        print("Faield at last")
        print(type(e))
        print(e)
    pass
